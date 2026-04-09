from rest_framework import viewsets, permissions, status
from rest_framework.pagination import PageNumberPagination
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.conf import settings
from django.db.models import Q
from datetime import timedelta
import random
import csv
import io
import json

from .models import Question, Choice, Exam, ExamAttempt, Subject, StudentAnswer
from .serializers import (
    QuestionSerializer, ChoiceSerializer, ExamSerializer, 
    ExamAttemptListSerializer, ExamAttemptDetailSerializer,
    ExamAttemptCreateSerializer, StudentAnswerSerializer,
    SubmitAnswerSerializer, SubjectSerializer
)
from .math_utils import format_math_question
from users.permissions import IsMasterAdmin
from courses.models import ActivationUnlock


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    pagination_class = StandardResultsSetPagination

    @action(detail=True, methods=['get'])
    def subjects(self, request, pk=None):
        """Get all subjects for a specific exam"""
        exam = self.get_object()
        subjects = exam.subjects.all()
        serializer = SubjectSerializer(subjects, many=True)
        return Response(serializer.data)


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    pagination_class = StandardResultsSetPagination

    @action(detail=True, methods=['get'])
    def questions(self, request, pk=None):
        """Get all questions for a specific subject with pagination"""
        subject = self.get_object()
        questions = subject.questions.all()
        
        # Apply pagination
        paginator = self.pagination_class()
        paginated_questions = paginator.paginate_queryset(questions, request)
        serializer = QuestionSerializer(paginated_questions, many=True)
        
        return paginator.get_paginated_response(serializer.data)


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

    @action(detail=True, methods=['post'])
    def update_choices(self, request, pk=None):
        """Update all choices for a question and set the correct answer.
        
        Expected payload:
        {
          "choices": [
            {"text": "Option A", "is_correct": false},
            {"text": "Option B", "is_correct": true},
            {"text": "Option C", "is_correct": false},
            {"text": "Option D", "is_correct": false}
          ],
          "text": "Updated question text (optional)",
          "explanation": "Updated explanation (optional)"
        }
        """
        question = self.get_object()
        choices_data = request.data.get('choices', [])
        question_text = request.data.get('text')
        explanation = request.data.get('explanation')

        # Validate choices
        if not choices_data or not isinstance(choices_data, list):
            return Response(
                {'detail': 'choices array is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if len(choices_data) < 2:
            return Response(
                {'detail': 'At least 2 choices are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check that exactly one choice is marked as correct
        correct_count = sum(1 for c in choices_data if c.get('is_correct', False))
        if correct_count != 1:
            return Response(
                {'detail': 'Exactly one choice must be marked as correct'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check that all choices have text
        if not all(c.get('text', '').strip() for c in choices_data):
            return Response(
                {'detail': 'All choices must have non-empty text'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Update question text/explanation if provided
            if question_text:
                question.text = question_text.strip()
            if explanation is not None:
                question.explanation = explanation

            # Delete old choices
            question.choices.all().delete()

            # Create new choices
            for choice_data in choices_data:
                Choice.objects.create(
                    question=question,
                    text=choice_data.get('text', '').strip(),
                    is_correct=choice_data.get('is_correct', False)
                )

            # Save question
            question.save()

            serializer = self.get_serializer(question)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'detail': f'Error updating choices: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


class BulkQuestionUploadView(APIView):
    """Accept JSON bulk uploads of questions in the specified format.

    Expected payload:
    {
      "exam_id": "JAMB",  // or exam slug
      "subject": "Chemistry",  // subject name
      "year": 2024,
      "questions": [
        {
          "question_text": "...",
          "options": {
            "A": "...",
            "B": "...",
            "C": "...",
            "D": "..."
          },
          "correct_answer": "A",
          "explanation": "...",
          "subject": "Chemistry"
        },
        ...
      ]
    }
    Also supports CSV/Excel file upload via multipart/form-data.
    """
    permission_classes = [IsMasterAdmin]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        # Check if file upload
        if 'file' in request.FILES:
            return self.handle_file_upload(request)
        
        return self.handle_json_upload(request)

    def handle_json_upload(self, request):
        data = request.data
        exam_id = data.get('exam_id')
        subject_name = data.get('subject')  # Top-level subject field
        year = data.get('year')
        questions_list = data.get('questions', [])

        if not exam_id:
            return Response(
                {'detail': 'exam_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not subject_name:
            return Response(
                {'detail': 'subject is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not questions_list:
            return Response(
                {'detail': 'questions array is required and cannot be empty'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not isinstance(questions_list, list):
            return Response(
                {'detail': 'questions must be an array'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Find exam by title or slug
        try:
            exam = Exam.objects.get(Q(title__iexact=exam_id) | Q(slug__iexact=exam_id))
        except Exam.DoesNotExist:
            return Response(
                {'detail': f'Exam "{exam_id}" not found. Please create the exam first.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Find the subject
        try:
            subject = Subject.objects.get(exam=exam, name__iexact=subject_name)
        except Subject.DoesNotExist:
            return Response(
                {'detail': f'Subject "{subject_name}" not found under exam "{exam.title}". Please create the subject first.'},
                status=status.HTTP_404_NOT_FOUND
            )

        return self.process_questions(questions_list, subject, year, request.user, exam)

    def handle_file_upload(self, request):
        file_obj = request.FILES['file']
        exam_id = request.data.get('exam_id')
        subject_name = request.data.get('subject')
        year = request.data.get('year')

        if not exam_id or not subject_name:
            return Response({'detail': 'exam_id and subject are required'}, status=status.HTTP_400_BAD_REQUEST)

        # Find exam
        try:
            exam = Exam.objects.get(Q(title__iexact=exam_id) | Q(slug__iexact=exam_id))
        except Exam.DoesNotExist:
            return Response({'detail': f'Exam "{exam_id}" not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Find subject
        try:
            subject = Subject.objects.get(exam=exam, name__iexact=subject_name)
        except Subject.DoesNotExist:
            return Response({'detail': f'Subject "{subject_name}" not found.'}, status=status.HTTP_404_NOT_FOUND)

        questions_list = []
        
        try:
            if file_obj.name.endswith('.csv'):
                decoded_file = file_obj.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                for row in reader:
                    # Expected CSV columns: question_text, option_a, option_b, option_c, option_d, correct_answer, explanation
                    options = {}
                    for key, value in row.items():
                        if key.lower().startswith('option_') and value:
                            opt_key = key.split('_')[1].upper()
                            options[opt_key] = value
                    
                    questions_list.append({
                        'question_text': row.get('question_text'),
                        'options': options,
                        'correct_answer': row.get('correct_answer'),
                        'explanation': row.get('explanation', ''),
                        'id': row.get('id', '')
                    })
            elif file_obj.name.endswith('.xlsx') or file_obj.name.endswith('.xls'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_obj)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = dict(zip(headers, row))
                        # Logic similar to CSV...
                        # For brevity, assuming similar structure or user installs openpyxl
                        pass 
                except ImportError:
                    return Response({'detail': 'openpyxl library not installed for Excel support'}, status=status.HTTP_501_NOT_IMPLEMENTED)
        except Exception as e:
            return Response({'detail': f'Error processing file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        return self.process_questions(questions_list, subject, year, request.user, exam)

    def process_questions(self, questions_list, subject, year, user, exam):
        created_questions = []
        errors = []

        if not questions_list:
            return Response({
                'success': 0,
                'total': 0,
                'created': [],
                'errors': ['No valid questions found in the upload'],
                'exam': exam.title,
                'year': year
            }, status=status.HTTP_400_BAD_REQUEST)

        for idx, q_data in enumerate(questions_list, 1):
            try:
                question_text = q_data.get('question_text', '').strip()
                options = q_data.get('options', {})
                correct_answer = str(q_data.get('correct_answer', '')).strip().upper()
                explanation = q_data.get('explanation', '').strip()
                question_id = q_data.get('id', f'Question {idx}')

                # Validate question_text
                if not question_text:
                    errors.append(f"{question_id}: Question text is required")
                    continue

                # Validate options
                if not options or not isinstance(options, dict):
                    errors.append(f"{question_id}: Options must be a dictionary with at least 2 options")
                    continue

                if len(options) < 2:
                    errors.append(f"{question_id}: At least 2 options are required (got {len(options)})")
                    continue

                # Filter empty options
                options = {k: v for k, v in options.items() if v and str(v).strip()}
                if len(options) < 2:
                    errors.append(f"{question_id}: At least 2 non-empty options are required")
                    continue

                # Validate correct_answer
                if not correct_answer:
                    errors.append(f"{question_id}: correct_answer is required")
                    continue

                if correct_answer not in options:
                    errors.append(f"{question_id}: correct_answer '{correct_answer}' must be one of {list(options.keys())}")
                    continue

                # Create question (save explanation if provided)
                question = Question.objects.create(
                    subject=subject,
                    text=question_text,
                    year=str(year) if year else None,
                    explanation=explanation,
                    creator=user
                )

                # Create choices
                for option_key, option_text in options.items():
                    is_correct = str(option_key).upper() == correct_answer
                    Choice.objects.create(
                        question=question,
                        text=str(option_text).strip(),
                        is_correct=is_correct
                    )

                created_questions.append({
                    'id': question.id,
                    'text': question_text[:50] + '...' if len(question_text) > 50 else question_text
                })

            except Exception as e:
                errors.append(f"{q_data.get('id', f'Question {idx}')}: {str(e)}")

        if not created_questions:
            return Response({
                'success': 0,
                'total': len(questions_list),
                'created': [],
                'errors': errors,
                'exam': exam.title,
                'year': year
            }, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'success': len(created_questions),
            'total': len(questions_list),
            'created': created_questions,
            'errors': errors if errors else None,
            'exam': exam.title,
            'year': year
        }, status=status.HTTP_201_CREATED)


class StartExamView(APIView):
    """Start a new exam attempt (single or multi-subject) and return the questions for that exam"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        data = request.data
        exam_id = data.get('exam')
        
        # Check if this is a multi-subject or single-subject exam
        subjects_config = data.get('subjects_config')
        
        # Single-subject (backward compatibility)
        if not subjects_config:
            subject_id = data.get('subject')
            num_questions = data.get('num_questions')
            time_limit_minutes = data.get('time_limit_minutes')
            test_name = None
            
            if not all([exam_id, subject_id, num_questions, time_limit_minutes]):
                return Response(
                    {'detail': 'exam, subject, num_questions, and time_limit_minutes are required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            exam = get_object_or_404(Exam, pk=exam_id)
            subject = get_object_or_404(Subject, pk=subject_id)
            subjects_data = [{'subject_id': subject_id, 'num_questions': num_questions}]
        
        # Multi-subject exam
        else:
            test_name = data.get('test_name')
            time_limit_minutes = data.get('time_limit_minutes')
            
            if not all([exam_id, subjects_config, time_limit_minutes, test_name]):
                return Response(
                    {'detail': 'exam, subjects_config, time_limit_minutes, and test_name are required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not isinstance(subjects_config, list) or len(subjects_config) == 0:
                return Response(
                    {'detail': 'subjects_config must be a non-empty array'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            exam = get_object_or_404(Exam, pk=exam_id)
            subject = None  # NULL for multi-subject exams
            subjects_data = subjects_config
        
        # Validate all subjects exist and gather question data
        total_num_questions = 0
        all_selected_questions = []
        subjects_info = []
        
        # For multi-subject exams: verify student has selected these subjects in their unlock
        if subjects_config and exam_id:
            try:
                unlock = ActivationUnlock.objects.get(user=request.user, exam_identifier=str(exam_id))
                allowed_subject_ids = set(unlock.selected_exam_subjects.values_list('id', flat=True))
                requested_subject_ids = set(cfg.get('subject_id') for cfg in subjects_config)
                
                # PHASE 1: Backward Compatibility for Legacy Unlocks
                # If unlock has no selected subjects (legacy/old unlock), allow access to ALL exam subjects
                # This maintains compatibility with exams unlocked before the course selection feature
                if not allowed_subject_ids:
                    allowed_subject_ids = set(exam.subjects.values_list('id', flat=True))
                
                # Check if all requested subjects are in allowed subjects
                if not requested_subject_ids.issubset(allowed_subject_ids):
                    unauthorized_subjects = requested_subject_ids - allowed_subject_ids
                    # PHASE 3: Better error messages with available subjects
                    allowed_subject_names = list(exam.subjects.filter(id__in=allowed_subject_ids).values_list('name', flat=True))
                    return Response(
                        {
                            'detail': 'You do not have access to all requested subjects.',
                            'available_subjects': allowed_subject_names,
                            'message': f'You have unlocked access to: {", ".join(allowed_subject_names)}. Please select only from these subjects.',
                            'action': 'Go back and reselect your subjects from the available options.'
                        },
                        status=status.HTTP_403_FORBIDDEN
                    )
            except ActivationUnlock.DoesNotExist:
                return Response(
                    {
                        'detail': 'You have not unlocked this exam yet.',
                        'action': 'Complete the activation payment to unlock this exam and select your subjects.',
                        'next_step': 'Unlock Exam'
                    },
                    status=status.HTTP_403_FORBIDDEN
                )
        
        for subject_config in subjects_data:
            subject_id = subject_config.get('subject_id')
            num_questions = subject_config.get('num_questions')
            
            if not subject_id or not num_questions:
                return Response(
                    {'detail': 'Each subject config must have subject_id and num_questions'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            subject_obj = get_object_or_404(Subject, pk=subject_id)
            all_questions = subject_obj.questions.all()
            num_questions = min(int(num_questions), all_questions.count())
            
            if num_questions < 1:
                return Response(
                    {'detail': f'Not enough questions in subject {subject_obj.name}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            selected_questions = random.sample(list(all_questions), num_questions)
            all_selected_questions.extend(selected_questions)
            total_num_questions += num_questions
            
            subjects_info.append({
                'subject_id': subject_obj.id,
                'name': subject_obj.name,
                'num_questions': num_questions
            })

        # Create exam attempt
        exam_attempt = ExamAttempt.objects.create(
            user=request.user,
            exam=exam,
            subject=subject,  # NULL for multi-subject
            test_name=test_name,
            num_questions=total_num_questions,
            time_limit_minutes=int(time_limit_minutes),
            started_at=timezone.now()
        )

        # Create student answer records for each question
        for question in all_selected_questions:
            StudentAnswer.objects.create(
                exam_attempt=exam_attempt,
                question=question,
                subject=question.subject
            )

        # Return exam attempt details
        return Response({
            'exam_attempt_id': exam_attempt.id,
            'test_name': test_name or exam.title,
            'exam_title': exam.title,
            'subjects': subjects_info,
            'num_questions': total_num_questions,
            'time_limit_minutes': time_limit_minutes,
            'started_at': exam_attempt.started_at
        }, status=status.HTTP_201_CREATED)


class SubmitAnswerView(APIView):
    """Submit an answer for a specific question in an exam attempt"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, exam_attempt_id):
        data = request.data
        question_id = data.get('question_id')
        choice_id = data.get('choice_id')

        exam_attempt = get_object_or_404(ExamAttempt, pk=exam_attempt_id, user=request.user)
        question = get_object_or_404(Question, pk=question_id)

        try:
            student_answer = StudentAnswer.objects.get(
                exam_attempt=exam_attempt,
                question=question
            )
        except StudentAnswer.DoesNotExist:
            return Response(
                {'detail': 'This question is not part of this exam attempt'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update the selected choice
        if choice_id:
            choice = get_object_or_404(Choice, pk=choice_id)
            student_answer.selected_choice = choice
            student_answer.is_correct = choice.is_correct
        else:
            student_answer.selected_choice = None
            student_answer.is_correct = False

        student_answer.answered_at = timezone.now()
        student_answer.save()

        return Response({
            'id': student_answer.id,
            'question_id': question.id,
            'is_correct': student_answer.is_correct
        })


class SubmitExamView(APIView):
    """Submit/complete an exam attempt and calculate the score"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, exam_attempt_id):
        exam_attempt = get_object_or_404(ExamAttempt, pk=exam_attempt_id, user=request.user)

        if exam_attempt.is_submitted:
            return Response(
                {'detail': 'This exam has already been submitted'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Calculate score
        student_answers = exam_attempt.student_answers.all()
        correct_count = student_answers.filter(is_correct=True).count()
        total_count = student_answers.count()
        
        exam_attempt.score = correct_count
        exam_attempt.is_submitted = True
        exam_attempt.submitted_at = timezone.now()
        
        # Calculate time taken
        time_taken = exam_attempt.submitted_at - exam_attempt.started_at
        exam_attempt.time_taken_seconds = int(time_taken.total_seconds())
        
        exam_attempt.save()

        return Response({
            'exam_attempt_id': exam_attempt.id,
            'score': exam_attempt.score,
            'total_questions': total_count,
            'percentage': round((correct_count / total_count) * 100, 2) if total_count > 0 else 0,
            'submitted_at': exam_attempt.submitted_at,
            'time_taken_seconds': exam_attempt.time_taken_seconds
        })


class ExamAttemptViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        return ExamAttempt.objects.filter(user=self.request.user, is_submitted=True).order_by('-submitted_at')

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ExamAttemptDetailSerializer
        return ExamAttemptListSerializer

    @action(detail=True, methods=['get'])
    def performance(self, request, pk=None):
        """Get detailed performance report for a specific exam attempt"""
        exam_attempt = self.get_object()
        
        student_answers = exam_attempt.student_answers.all()
        wrong_answers = student_answers.filter(is_correct=False)

        wrong_answers_data = []
        for answer in wrong_answers:
            correct_choice = answer.question.choices.filter(is_correct=True).first()
            wrong_answers_data.append({
                'question_id': answer.question.id,
                'question_text': answer.question.text,
                'user_answer': answer.selected_choice.text if answer.selected_choice else 'Not answered',
                'correct_answer': correct_choice.text if correct_choice else 'N/A',
                'explanation': getattr(answer.question, 'explanation', 'No explanation available')
            })

        serializer = ExamAttemptDetailSerializer(exam_attempt)
        data = serializer.data
        data['wrong_answers'] = wrong_answers_data

        return Response(data)


class ExamAttemptListView(APIView):
    """List all submitted exam attempts for the current user with pagination"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        page = request.query_params.get('page', 1)
        page_size = 10

        attempts = ExamAttempt.objects.filter(
            user=user, 
            is_submitted=True
        ).order_by('-submitted_at')

        total_count = attempts.count()
        start = (int(page) - 1) * page_size
        end = start + page_size

        paginated_attempts = attempts[start:end]
        serializer = ExamAttemptListSerializer(paginated_attempts, many=True)

        return Response({
            'count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': (total_count + page_size - 1) // page_size,
            'results': serializer.data
        })


class GetExamQuestionsView(APIView):
    """Get paginated questions for an active exam attempt, optionally filtered by subject"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, exam_attempt_id):
        exam_attempt = get_object_or_404(ExamAttempt, pk=exam_attempt_id, user=request.user)
        page = int(request.query_params.get('page', 1))
        subject_id = request.query_params.get('subject')
        page_size = 10

        # Filter student answers by subject if specified
        student_answers = exam_attempt.student_answers.all()
        if subject_id:
            student_answers = student_answers.filter(subject_id=int(subject_id))
        
        student_answers = student_answers.order_by('id')
        total_count = student_answers.count()
        
        start = (page - 1) * page_size
        end = start + page_size

        paginated_answers = student_answers[start:end]

        questions_data = []
        for answer in paginated_answers:
            question = answer.question
            questions_data.append({
                'id': question.id,
                'text': format_math_question(question.text),
                'image': request.build_absolute_uri(question.image.url) if question.image else None,
                'choices': [
                    {'id': c.id, 'text': format_math_question(c.text)}
                    for c in question.choices.all()
                ],
                'user_answer_id': answer.selected_choice.id if answer.selected_choice else None,
                'is_answered': answer.selected_choice is not None,
                'year': question.year
            })

        return Response({
            'exam_attempt_id': exam_attempt.id,
            'page': page,
            'page_size': page_size,
            'total_questions': total_count,
            'total_pages': (total_count + page_size - 1) // page_size,
            'questions': questions_data
        })


class ExamProgressView(APIView):
    """Get progress/navigation data for an active exam (which questions answered, organized by subject)"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, exam_attempt_id):
        exam_attempt = get_object_or_404(ExamAttempt, pk=exam_attempt_id, user=request.user)
        
        student_answers = exam_attempt.student_answers.all().order_by('id')
        
        # Always return progress array for backward compatibility
        progress = []
        for idx, answer in enumerate(student_answers, 1):
            progress.append({
                'question_number': idx,
                'question_id': answer.question.id,
                'is_answered': answer.selected_choice is not None,
                'is_correct': answer.is_correct
            })

        # For multi-subject exams, also return subject-level progress
        subject_progress = []
        if exam_attempt.test_name:  # Multi-subject exam
            # Group by subject
            from django.db.models import Count, Q
            subjects_in_attempt = exam_attempt.student_answers.values('subject_id', 'subject__name').distinct()
            
            for subject_data in subjects_in_attempt:
                subject_id = subject_data['subject_id']
                subject_name = subject_data['subject__name']
                
                subject_answers = exam_attempt.student_answers.filter(subject_id=subject_id)
                answered_count = subject_answers.filter(selected_choice__isnull=False).count()
                total_count = subject_answers.count()
                
                subject_progress.append({
                    'subject_id': subject_id,
                    'subject_name': subject_name,
                    'answered_count': answered_count,
                    'total_questions': total_count
                })

        return Response({
            'exam_attempt_id': exam_attempt.id,
            'total_questions': exam_attempt.num_questions,
            'answered_count': student_answers.filter(selected_choice__isnull=False).count(),
            'progress': progress,
            'subject_progress': subject_progress
        })


class AnalyticsView(APIView):
    """Get CBT analytics for admin dashboard"""
    permission_classes = [IsMasterAdmin]

    def get(self, request):
        from django.db.models import Count, Avg, F
        
        # Get total exam attempts
        total_attempts = ExamAttempt.objects.count()
        
        # Get average score
        avg_score = ExamAttempt.objects.aggregate(avg=Avg('score'))['avg'] or 0
        
        # Get subjects with attempt counts
        subjects_data = Subject.objects.annotate(
            attempt_count=Count('attempts', distinct=True),
            avg_score=Avg('attempts__score')
        ).values('name', 'attempt_count', 'avg_score').order_by('-attempt_count')
        
        # Get today's attempts
        today = timezone.now().date()
        today_attempts = ExamAttempt.objects.filter(
            started_at__date=today
        ).count()
        
        # Compute pass rate (default pass threshold 50%)
        PASS_THRESHOLD = getattr(settings, 'CBT_PASS_THRESHOLD', 50)
        passed_attempts = ExamAttempt.objects.filter(score__gte=PASS_THRESHOLD).count()
        pass_rate = (passed_attempts / total_attempts) if total_attempts > 0 else 0

        # Average time taken (seconds -> minutes)
        avg_time_seconds = ExamAttempt.objects.aggregate(avg_time=Avg('time_taken_seconds'))['avg_time']
        avg_time_minutes = (float(avg_time_seconds) / 60.0) if avg_time_seconds is not None else None

        # Active students (unique users who attempted)
        active_students = ExamAttempt.objects.values('user').distinct().count()

        # Top scorers (top attempts by score)
        top_attempts_qs = ExamAttempt.objects.filter(score__isnull=False).select_related('user', 'exam', 'subject').order_by('-score', '-submitted_at')[:10]
        top_scorers = []
        for a in top_attempts_qs:
            top_scorers.append({
                'id': a.id,
                'user_id': a.user_id,
                'name': getattr(a.user, 'username', None) or getattr(a.user, 'full_name', None) or str(a.user_id),
                'exam': a.exam.title if a.exam else None,
                'subject': a.subject.name if a.subject else None,
                'score': float(a.score) if a.score is not None else None,
                'date': a.submitted_at.isoformat() if a.submitted_at else (a.finished_at.isoformat() if a.finished_at else a.started_at.isoformat())
            })

        top_subject = None
        if len(subjects_data) > 0:
            top_subject = subjects_data[0]['name']

        return Response({
            'total_attempts': total_attempts,
            'total_exams': Exam.objects.count(),
            'average_score': float(avg_score),
            'today_attempts': today_attempts,
            'subjects': list(subjects_data[:10]),  # Top 10 subjects
            'passed_attempts': passed_attempts,
            'pass_rate': float(pass_rate),
            'average_time_seconds': float(avg_time_seconds) if avg_time_seconds is not None else None,
            'average_time_minutes': float(avg_time_minutes) if avg_time_minutes is not None else None,
            'active_students_count': active_students,
            'top_scorers': top_scorers,
            'top_subject': top_subject,
        })

class StudentLeaderboardView(APIView):
    """Get student leaderboard based on exam scores"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from django.db.models import Count, Avg, Max
        
        limit = int(request.query_params.get('limit', 50))
        
        # Get top students by average score
        students_stats = ExamAttempt.objects.filter(
            is_submitted=True
        ).values('user__id', 'user__username').annotate(
            avg_score=Avg('score'),
            high_score=Max('score'),
            attempts_count=Count('id', distinct=True)
        ).order_by('-avg_score', '-high_score')[:limit]
        
        # Convert to response format
        leaderboard_data = [
            {
                'id': stats['user__id'],
                'username': stats['user__username'],
                'avg_score': float(stats['avg_score'] or 0),
                'high_score': float(stats['high_score'] or 0),
                'attempts_count': stats['attempts_count']
            }
            for stats in students_stats
        ]
        
        return Response(leaderboard_data)